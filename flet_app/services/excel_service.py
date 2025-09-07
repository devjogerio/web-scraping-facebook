"""Serviço de exportação para Excel - Versão Flet.

Este módulo implementa a lógica de exportação de dados
extraídos para arquivos Excel formatados.
"""

import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from flet_app.config.logging_config import get_logger
from flet_app.models.facebook_data import FacebookData
from flet_app.models.export_job import ExportJob


class ExcelService:
    """Serviço para exportação de dados para Excel.
    
    Implementa métodos para gerar arquivos Excel formatados
    com os dados extraídos do Facebook.
    """
    
    def __init__(self, export_directory: str = 'exports'):
        """Inicializar serviço de exportação.
        
        Args:
            export_directory: Diretório para salvar arquivos exportados
        """
        self.export_directory = export_directory
        self.logger = get_logger('ExcelService')
        
        # Criar diretório de exportação se não existir
        if not os.path.exists(export_directory):
            os.makedirs(export_directory)
            self.logger.info(f'Diretório de exportação criado: {export_directory}')
    
    def export_task_data(self, task_name: str, data_list: List[FacebookData], 
                        task_config: Dict[str, Any] = None) -> tuple[str, str, int]:
        """Exportar dados de uma tarefa para Excel.
        
        Args:
            task_name: Nome da tarefa
            data_list: Lista de dados extraídos
            task_config: Configurações da tarefa
            
        Returns:
            Tupla (nome_arquivo, caminho_completo, tamanho_arquivo)
        """
        try:
            # Gerar nome do arquivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_task_name = self._sanitize_filename(task_name)
            filename = f'{safe_task_name}_{timestamp}.xlsx'
            file_path = os.path.join(self.export_directory, filename)
            
            self.logger.info(f'Iniciando exportação para: {filename}')
            
            # Criar workbook
            wb = Workbook()
            
            # Remover planilha padrão
            wb.remove(wb.active)
            
            # Agrupar dados por tipo
            data_by_type = self._group_data_by_type(data_list)
            
            # Criar planilha de resumo
            self._create_summary_sheet(wb, task_name, data_by_type, task_config)
            
            # Criar planilhas para cada tipo de dado
            for data_type, data_items in data_by_type.items():
                if data_items:
                    self._create_data_sheet(wb, data_type, data_items)
            
            # Salvar arquivo
            wb.save(file_path)
            
            # Obter tamanho do arquivo
            file_size = os.path.getsize(file_path)
            
            self.logger.info(f'Exportação concluída: {filename} ({self._format_file_size(file_size)})')
            
            return filename, file_path, file_size
            
        except Exception as e:
            self.logger.error(f'Erro durante exportação: {str(e)}')
            raise
    
    def _group_data_by_type(self, data_list: List[FacebookData]) -> Dict[str, List[FacebookData]]:
        """Agrupar dados por tipo.
        
        Args:
            data_list: Lista de dados
            
        Returns:
            Dicionário agrupado por tipo
        """
        grouped = {}
        
        for data in data_list:
            data_type = data.data_type
            if data_type not in grouped:
                grouped[data_type] = []
            grouped[data_type].append(data)
        
        return grouped
    
    def _create_summary_sheet(self, wb: Workbook, task_name: str, 
                             data_by_type: Dict[str, List[FacebookData]], 
                             task_config: Dict[str, Any] = None) -> None:
        """Criar planilha de resumo.
        
        Args:
            wb: Workbook do Excel
            task_name: Nome da tarefa
            data_by_type: Dados agrupados por tipo
            task_config: Configurações da tarefa
        """
        ws = wb.create_sheet('Resumo', 0)
        
        # Título
        ws['A1'] = 'Relatório de Extração - Facebook'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Informações da tarefa
        row = 3
        ws[f'A{row}'] = 'Nome da Tarefa:'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = task_name
        
        row += 1
        ws[f'A{row}'] = 'Data de Exportação:'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
        if task_config:
            row += 1
            ws[f'A{row}'] = 'URL:'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = task_config.get('url', 'N/A')
        
        # Estatísticas
        row += 3
        ws[f'A{row}'] = 'Estatísticas de Extração'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        row += 2
        ws[f'A{row}'] = 'Tipo de Dado'
        ws[f'B{row}'] = 'Quantidade'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        total_items = 0
        for data_type, data_items in data_by_type.items():
            row += 1
            ws[f'A{row}'] = data_type.title()
            ws[f'B{row}'] = len(data_items)
            total_items += len(data_items)
        
        row += 1
        ws[f'A{row}'] = 'Total'
        ws[f'B{row}'] = total_items
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_data_sheet(self, wb: Workbook, data_type: str, data_items: List[FacebookData]) -> None:
        """Criar planilha para um tipo específico de dados.
        
        Args:
            wb: Workbook do Excel
            data_type: Tipo de dados
            data_items: Lista de dados do tipo
        """
        sheet_name = data_type.title()
        ws = wb.create_sheet(sheet_name)
        
        # Converter dados para DataFrame
        data_rows = [item.to_excel_row() for item in data_items]
        
        if not data_rows:
            ws['A1'] = f'Nenhum dado do tipo {data_type} encontrado'
            return
        
        df = pd.DataFrame(data_rows)
        
        # Adicionar dados à planilha
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Formatação do cabeçalho
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limitar largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _sanitize_filename(self, filename: str) -> str:
        """Sanitizar nome do arquivo removendo caracteres inválidos.
        
        Args:
            filename: Nome original do arquivo
            
        Returns:
            Nome sanitizado
        """
        import re
        
        # Remover caracteres inválidos
        sanitized = re.sub(r'[<>:"/\\|?*]', '_', filename)
        
        # Limitar comprimento
        if len(sanitized) > 50:
            sanitized = sanitized[:50]
        
        return sanitized
    
    def _format_file_size(self, size_bytes: int) -> str:
        """Formatar tamanho do arquivo.
        
        Args:
            size_bytes: Tamanho em bytes
            
        Returns:
            Tamanho formatado
        """
        if size_bytes < 1024:
            return f'{size_bytes} B'
        elif size_bytes < 1024 * 1024:
            return f'{size_bytes / 1024:.1f} KB'
        elif size_bytes < 1024 * 1024 * 1024:
            return f'{size_bytes / (1024 * 1024):.1f} MB'
        else:
            return f'{size_bytes / (1024 * 1024 * 1024):.1f} GB'
    
    def create_template_file(self, filename: str = 'template.xlsx') -> str:
        """Criar arquivo template para demonstração.
        
        Args:
            filename: Nome do arquivo template
            
        Returns:
            Caminho do arquivo criado
        """
        try:
            file_path = os.path.join(self.export_directory, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Template'
            
            # Cabeçalhos de exemplo
            headers = ['ID', 'Tipo', 'Conteúdo', 'Autor', 'Data/Hora', 'Curtidas', 'Comentários']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Dados de exemplo
            example_data = [
                ['1', 'Post', 'Exemplo de post do Facebook', 'Usuário Exemplo', '01/01/2024 10:00', '10', '5'],
                ['2', 'Comment', 'Exemplo de comentário', 'Outro Usuário', '01/01/2024 10:05', '2', '0']
            ]
            
            for row_idx, row_data in enumerate(example_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            self.logger.info(f'Arquivo template criado: {filename}')
            
            return file_path
            
        except Exception as e:
            self.logger.error(f'Erro ao criar arquivo template: {str(e)}')
            raise
    
    def validate_export_directory(self) -> bool:
        """Validar se o diretório de exportação existe e é acessível.
        
        Returns:
            True se o diretório é válido
        """
        try:
            if not os.path.exists(self.export_directory):
                os.makedirs(self.export_directory)
            
            # Testar escrita no diretório
            test_file = os.path.join(self.export_directory, '.test')
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            
            return True
            
        except Exception as e:
            self.logger.error(f'Erro ao validar diretório de exportação: {str(e)}')
            return False
    
    def get_export_directory(self) -> str:
        """Obter caminho do diretório de exportação.
        
        Returns:
            Caminho do diretório
        """
        return os.path.abspath(self.export_directory)
    
    def list_exported_files(self) -> List[Dict[str, Any]]:
        """Listar arquivos exportados no diretório.
        
        Returns:
            Lista de informações dos arquivos
        """
        files = []
        
        try:
            if not os.path.exists(self.export_directory):
                return files
            
            for filename in os.listdir(self.export_directory):
                if filename.endswith('.xlsx'):
                    file_path = os.path.join(self.export_directory, filename)
                    file_stat = os.stat(file_path)
                    
                    files.append({
                        'filename': filename,
                        'path': file_path,
                        'size': file_stat.st_size,
                        'size_formatted': self._format_file_size(file_stat.st_size),
                        'created_at': datetime.fromtimestamp(file_stat.st_ctime),
                        'modified_at': datetime.fromtimestamp(file_stat.st_mtime)
                    })
            
            # Ordenar por data de modificação (mais recente primeiro)
            files.sort(key=lambda x: x['modified_at'], reverse=True)
            
        except Exception as e:
            self.logger.error(f'Erro ao listar arquivos exportados: {str(e)}')
        
        return files