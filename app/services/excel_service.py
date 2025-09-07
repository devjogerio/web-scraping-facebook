"""Serviço de geração de arquivos Excel.

Este módulo implementa a lógica para criação de arquivos Excel
formatados com os dados extraídos do Facebook.
"""

import os
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelService:
    """Serviço para geração de arquivos Excel.
    
    Implementa métodos para criar arquivos Excel formatados
    com múltiplas planilhas e estilos profissionais.
    """
    
    def __init__(self):
        """Inicializar serviço de Excel."""
        self.logger = logging.getLogger(__name__)
        
        # Estilos padrão
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1877F2", end_color="1877F2", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_excel_file(self, file_path: str, task_info: Dict[str, Any], 
                         organized_data: Dict[str, List[Dict[str, Any]]], 
                         options: Dict[str, Any]) -> None:
        """Criar arquivo Excel com os dados organizados.
        
        Args:
            file_path: Caminho do arquivo a ser criado
            task_info: Informações da tarefa de scraping
            organized_data: Dados organizados por tipo
            options: Opções de formatação
        """
        try:
            self.logger.info(f"Criando arquivo Excel: {file_path}")
            
            # Criar workbook
            wb = Workbook()
            
            # Remover planilha padrão
            wb.remove(wb.active)
            
            # Criar planilha de resumo primeiro
            if options.get('include_summary', True):
                self._create_summary_sheet(wb, task_info, organized_data, options)
            
            # Criar planilhas para cada tipo de dado
            if options.get('separate_sheets', True):
                self._create_data_sheets(wb, organized_data, options)
            else:
                self._create_combined_sheet(wb, organized_data, options)
            
            # Salvar arquivo
            wb.save(file_path)
            
            self.logger.info(f"Arquivo Excel criado com sucesso: {file_path}")
            
        except Exception as e:
            self.logger.error(f"Erro ao criar arquivo Excel: {str(e)}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, task_info: Dict[str, Any], 
                             organized_data: Dict[str, List], options: Dict[str, Any]) -> None:
        """Criar planilha de resumo.
        
        Args:
            wb: Workbook do Excel
            task_info: Informações da tarefa
            organized_data: Dados organizados
            options: Opções de formatação
        """
        ws = wb.create_sheet("Resumo", 0)
        
        # Título principal
        ws['A1'] = "Relatório de Scraping do Facebook"
        ws['A1'].font = Font(size=16, bold=True, color="1877F2")
        ws.merge_cells('A1:D1')
        
        # Informações da tarefa
        row = 3
        task_fields = [
            ('Nome da Tarefa:', task_info.get('name', 'N/A')),
            ('URL:', task_info.get('url', 'N/A')),
            ('Status:', task_info.get('status', 'N/A')),
            ('Data de Criação:', self._format_datetime(task_info.get('created_at'))),
            ('Data de Conclusão:', self._format_datetime(task_info.get('completed_at'))),
            ('Itens Processados:', str(task_info.get('items_processed', 0))),
            ('Duração:', self._format_duration(task_info.get('duration')))
        ]
        
        for label, value in task_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Estatísticas por tipo
        row += 2
        ws[f'A{row}'] = "Estatísticas por Tipo de Dado"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Cabeçalhos da tabela
        headers = ['Tipo de Dado', 'Quantidade', 'Percentual']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        row += 1
        
        # Dados da tabela
        total_items = sum(len(items) for key, items in organized_data.items() if key != 'summary')
        
        for data_type, items in organized_data.items():
            if data_type != 'summary' and items:
                count = len(items)
                percentage = (count / total_items * 100) if total_items > 0 else 0
                
                ws.cell(row=row, column=1, value=data_type.title()).border = self.border
                ws.cell(row=row, column=2, value=count).border = self.border
                ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = self.border
                row += 1
        
        # Total
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=1).border = self.border
        ws.cell(row=row, column=2, value=total_items).font = Font(bold=True)
        ws.cell(row=row, column=2).border = self.border
        ws.cell(row=row, column=3, value="100.0%").font = Font(bold=True)
        ws.cell(row=row, column=3).border = self.border
        
        # Criar gráfico se solicitado
        if options.get('include_charts', False):
            self._create_summary_chart(ws, organized_data, row + 2)
        
        # Ajustar largura das colunas
        if options.get('auto_fit_columns', True):
            self._auto_fit_columns(ws)
    
    def _create_data_sheets(self, wb: Workbook, organized_data: Dict[str, List], 
                           options: Dict[str, Any]) -> None:
        """Criar planilhas separadas para cada tipo de dado.
        
        Args:
            wb: Workbook do Excel
            organized_data: Dados organizados por tipo
            options: Opções de formatação
        """
        for data_type, items in organized_data.items():
            if data_type != 'summary' and items:
                sheet_name = data_type.title()
                ws = wb.create_sheet(sheet_name)
                
                # Converter para DataFrame
                df = pd.DataFrame(items)
                
                # Limitar tamanho do conteúdo se especificado
                max_length = options.get('max_content_length', 1000)
                if 'Conteúdo' in df.columns:
                    df['Conteúdo'] = df['Conteúdo'].astype(str).str[:max_length]
                
                # Adicionar dados à planilha
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # Aplicar formatação
                if options.get('apply_formatting', True):
                    self._apply_sheet_formatting(ws, len(df.columns), len(df) + 1, options)
    
    def _create_combined_sheet(self, wb: Workbook, organized_data: Dict[str, List], 
                              options: Dict[str, Any]) -> None:
        """Criar planilha única com todos os dados.
        
        Args:
            wb: Workbook do Excel
            organized_data: Dados organizados por tipo
            options: Opções de formatação
        """
        ws = wb.create_sheet("Todos os Dados")
        
        # Combinar todos os dados
        all_data = []
        for data_type, items in organized_data.items():
            if data_type != 'summary':
                for item in items:
                    item_copy = item.copy()
                    item_copy['Tipo de Dado'] = data_type.title()
                    all_data.append(item_copy)
        
        if all_data:
            # Converter para DataFrame
            df = pd.DataFrame(all_data)
            
            # Reordenar colunas para colocar 'Tipo de Dado' primeiro
            if 'Tipo de Dado' in df.columns:
                cols = ['Tipo de Dado'] + [col for col in df.columns if col != 'Tipo de Dado']
                df = df[cols]
            
            # Limitar tamanho do conteúdo
            max_length = options.get('max_content_length', 1000)
            if 'Conteúdo' in df.columns:
                df['Conteúdo'] = df['Conteúdo'].astype(str).str[:max_length]
            
            # Adicionar dados à planilha
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Aplicar formatação
            if options.get('apply_formatting', True):
                self._apply_sheet_formatting(ws, len(df.columns), len(df) + 1, options)
    
    def _apply_sheet_formatting(self, ws, num_cols: int, num_rows: int, 
                               options: Dict[str, Any]) -> None:
        """Aplicar formatação a uma planilha.
        
        Args:
            ws: Worksheet
            num_cols: Número de colunas
            num_rows: Número de linhas
            options: Opções de formatação
        """
        # Formatação do cabeçalho
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Formatação dos dados
        for row in range(2, num_rows + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.alignment = self.data_alignment
                cell.border = self.border
        
        # Congelar primeira linha se solicitado
        if options.get('freeze_header_row', True):
            ws.freeze_panes = 'A2'
        
        # Ajustar largura das colunas
        if options.get('auto_fit_columns', True):
            self._auto_fit_columns(ws)
        
        # Criar tabela formatada
        if num_rows > 1:
            table_range = f"A1:{self._get_column_letter(num_cols)}{num_rows}"
            table = Table(displayName=f"Table_{ws.title.replace(' ', '_')}", ref=table_range)
            
            # Estilo da tabela
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            try:
                ws.add_table(table)
            except Exception as e:
                self.logger.warning(f"Erro ao adicionar tabela: {str(e)}")
    
    def _create_summary_chart(self, ws, organized_data: Dict[str, List], start_row: int) -> None:
        """Criar gráfico de resumo.
        
        Args:
            ws: Worksheet
            organized_data: Dados organizados
            start_row: Linha inicial para o gráfico
        """
        try:
            # Preparar dados para o gráfico
            chart_data = []
            for data_type, items in organized_data.items():
                if data_type != 'summary' and items:
                    chart_data.append((data_type.title(), len(items)))
            
            if not chart_data:
                return
            
            # Criar dados do gráfico na planilha
            chart_start_row = start_row
            ws.cell(row=chart_start_row, column=1, value="Tipo")
            ws.cell(row=chart_start_row, column=2, value="Quantidade")
            
            for i, (data_type, count) in enumerate(chart_data, 1):
                ws.cell(row=chart_start_row + i, column=1, value=data_type)
                ws.cell(row=chart_start_row + i, column=2, value=count)
            
            # Criar gráfico de barras
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Distribuição dos Dados Extraídos"
            chart.y_axis.title = 'Quantidade'
            chart.x_axis.title = 'Tipo de Dado'
            
            # Definir dados do gráfico
            data = Reference(ws, min_col=2, min_row=chart_start_row, 
                           max_row=chart_start_row + len(chart_data))
            cats = Reference(ws, min_col=1, min_row=chart_start_row + 1, 
                           max_row=chart_start_row + len(chart_data))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Adicionar gráfico à planilha
            ws.add_chart(chart, f"D{start_row}")
            
        except Exception as e:
            self.logger.warning(f"Erro ao criar gráfico: {str(e)}")
    
    def _auto_fit_columns(self, ws) -> None:
        """Ajustar automaticamente a largura das colunas.
        
        Args:
            ws: Worksheet
        """
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Definir largura mínima e máxima
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"Erro ao ajustar largura das colunas: {str(e)}")
    
    def _get_column_letter(self, col_num: int) -> str:
        """Converter número da coluna para letra.
        
        Args:
            col_num: Número da coluna
            
        Returns:
            Letra da coluna
        """
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_num)
    
    def _format_datetime(self, dt_string: Optional[str]) -> str:
        """Formatar string de datetime.
        
        Args:
            dt_string: String de datetime ISO
            
        Returns:
            Datetime formatado ou 'N/A'
        """
        if not dt_string:
            return 'N/A'
        
        try:
            dt = datetime.fromisoformat(dt_string.replace('Z', '+00:00'))
            return dt.strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            return dt_string
    
    def _format_duration(self, duration: Optional[float]) -> str:
        """Formatar duração em segundos.
        
        Args:
            duration: Duração em segundos
            
        Returns:
            Duração formatada ou 'N/A'
        """
        if duration is None:
            return 'N/A'
        
        try:
            hours = int(duration // 3600)
            minutes = int((duration % 3600) // 60)
            seconds = int(duration % 60)
            
            if hours > 0:
                return f"{hours}h {minutes}m {seconds}s"
            elif minutes > 0:
                return f"{minutes}m {seconds}s"
            else:
                return f"{seconds}s"
        except Exception:
            return str(duration)
    
    def validate_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Validar arquivo Excel criado.
        
        Args:
            file_path: Caminho do arquivo
            
        Returns:
            Resultado da validação
        """
        try:
            if not os.path.exists(file_path):
                return {'valid': False, 'error': 'Arquivo não encontrado'}
            
            # Verificar se é um arquivo Excel válido
            wb = Workbook()
            wb.load_workbook(file_path)
            
            # Verificar tamanho do arquivo
            file_size = os.path.getsize(file_path)
            
            return {
                'valid': True,
                'file_size': file_size,
                'sheets_count': len(wb.sheetnames),
                'sheets': wb.sheetnames
            }
            
        except Exception as e:
            return {'valid': False, 'error': str(e)}